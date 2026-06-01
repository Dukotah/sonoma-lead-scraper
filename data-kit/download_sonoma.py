"""
Downloads Overture Maps Places data for Sonoma County, CA.
Outputs three files:
  - sonoma_places.parquet  (raw Overture data, ~20 columns, all types)
  - sonoma_places.csv      (same data, flattened, Excel-friendly)
  - sonoma_businesses.xlsx (Excel summary: chains removed, tiered by lead quality)

Run:  python download_sonoma.py
First run takes 2-5 minutes (streams partitioned parquet from S3 and filters to the bbox).
"""
import sys, subprocess

# Windows consoles default to cp1252, which can't encode characters like the arrow
# below and crashes on print. Force UTF-8 output (replace anything unencodable).
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

def pip_install(*pkgs):
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--quiet", *pkgs])

try:
    import duckdb
except ImportError:
    print("Installing duckdb...")
    pip_install("duckdb")
    import duckdb

try:
    import openpyxl  # noqa
except ImportError:
    print("Installing openpyxl...")
    pip_install("openpyxl")

def _get_fsspec():
    """Lazily import fsspec/s3fs, installing s3fs if needed."""
    try:
        import fsspec
    except ImportError:
        print("Installing s3fs...")
        pip_install("s3fs")
        import fsspec
    return fsspec

# Sonoma County bounding box
BBOX = {"south": 38.05, "west": -123.55, "north": 38.85, "east": -122.35}

# Known-good release used ONLY as a fallback. The script auto-detects the newest
# release from S3 on every run (see detect_latest_release), so this rarely matters
# — it just keeps the script working if the lookup ever fails (offline, schema change).
FALLBACK_RELEASE = "2026-05-20.0"

# A business name that appears at least this many times across the county is almost
# always a chain/franchise or a non-business POI (ATM, post office) — not a web-design
# lead. This backstops Overture's `brand` field, which is only ~half-populated.
CHAIN_NAME_THRESHOLD = 5

# Overture's "places" theme includes natural features and civic POIs that nobody
# buys a website for. Drop these categories from the lead summary (they stay in the
# raw .parquet/.csv). Add or remove entries to taste for a different region.
NON_BUSINESS_CATEGORIES = {
    "park", "beach", "river", "mountain", "mountain_peak", "hiking_trail", "trail",
    "lake", "forest", "island", "dam", "waterfall", "canyon", "valley", "cliff",
    "plateau", "natural_feature", "structure_and_geography", "body_of_water",
    "landmark_and_historical_building", "monument", "bridge", "cemetery",
    "post_office", "fire_station", "police_department",
}

# Domains that mean "no real website" — a social/listing page, not an owned site.
# Distinctive substrings only (avoid e.g. bare "x.com" which would match fedex.com).
WEAK_DOMAINS = [
    "facebook.com", "instagram.com", "yelp.com", "linktr.ee", "linktree.com",
    "yellowpages.com", "business.site", "wixsite.com", "tiktok.com",
    "linkedin.com", "twitter.com", "nextdoor.com", "google.com/maps",
]

print("Connecting to DuckDB + Overture S3...")
con = duckdb.connect()

# Primary path: DuckDB's httpfs extension talks to S3 directly. It's the fastest
# option, but it must download the extension binary from extensions.duckdb.org on
# first use. Locked-down networks (corporate proxies, sandboxed CI, Claude Code on
# the web) often allow the S3 data host but block the extension host. When that
# happens we fall back to reading the same public bucket through fsspec/s3fs, which
# only needs the S3 host itself. DuckDB routes s3:// reads through any registered
# fsspec filesystem, so the rest of the script is unchanged.
S3FS = None
try:
    con.execute("INSTALL httpfs; LOAD httpfs;")
    con.execute("SET s3_region='us-west-2';")
    print("  connection: DuckDB httpfs extension")
except Exception as e:
    print(f"  httpfs unavailable ({type(e).__name__}); falling back to fsspec/s3fs")
    fsspec = _get_fsspec()
    S3FS = fsspec.filesystem("s3", anon=True, client_kwargs={"region_name": "us-west-2"})
    con.register_filesystem(S3FS)
    print("  connection: fsspec/s3fs (anonymous)")


def detect_latest_release() -> str:
    """Return the newest Overture release string from S3.
    Falls back to FALLBACK_RELEASE if the lookup fails for any reason."""
    try:
        if S3FS is not None:
            rels = [p.rstrip("/").split("/")[-1]
                    for p in S3FS.ls("overturemaps-us-west-2/release")]
            rels = [r for r in rels if r[:3].isdigit() or r.startswith("202")]
            if rels:
                return sorted(rels)[-1]
        else:
            pat = ("s3://overturemaps-us-west-2/release/202[0-9]-*"
                   "/theme=places/type=place/part-00000-*")
            rows = con.execute(
                "SELECT DISTINCT regexp_extract(file, 'release/([^/]+)/', 1) AS rel "
                f"FROM glob('{pat}') WHERE rel <> '' ORDER BY rel DESC LIMIT 1"
            ).fetchall()
            if rows and rows[0][0]:
                return rows[0][0]
    except Exception as e:
        print(f"  (release auto-detect failed: {type(e).__name__}; using fallback {FALLBACK_RELEASE})")
    return FALLBACK_RELEASE


RELEASE = detect_latest_release()
print(f"Using Overture release: {RELEASE}")

S3_URL = (
    f"s3://overturemaps-us-west-2/release/{RELEASE}"
    f"/theme=places/type=place/*"
)

print(f"Querying Overture for Sonoma County...")
print("(2-5 min: streaming partitioned parquet from S3)")

# Pull a clean set of useful columns
query = f"""
COPY (
  SELECT
    id,
    names.primary AS name,
    categories.primary AS category_primary,
    categories.alternate AS category_alt,
    confidence,
    websites,
    socials,
    emails,
    phones,
    brand.names.primary AS brand,
    addresses[1].freeform AS address_line,
    addresses[1].locality AS city,
    addresses[1].region   AS state,
    addresses[1].postcode AS zip,
    addresses[1].country  AS country,
    bbox.xmin AS lon,
    bbox.ymin AS lat,
    sources[1].dataset AS source_dataset,
    sources[1].record_id AS source_id
  FROM read_parquet('{S3_URL}', hive_partitioning=1)
  WHERE bbox.xmin BETWEEN {BBOX["west"]} AND {BBOX["east"]}
    AND bbox.ymin BETWEEN {BBOX["south"]} AND {BBOX["north"]}
) TO 'sonoma_places.parquet' (FORMAT PARQUET);
"""
con.execute(query)
n = con.execute("SELECT COUNT(*) FROM 'sonoma_places.parquet'").fetchone()[0]
print(f"  -> wrote sonoma_places.parquet ({n:,} records)")

# Flatten arrays for CSV/xlsx readability
print("Flattening for CSV/xlsx...")
con.execute("""
CREATE TABLE flat AS
SELECT
  id, name, category_primary,
  array_to_string(category_alt, '|') AS category_alt,
  confidence,
  CASE WHEN length(websites) > 0 THEN websites[1] ELSE NULL END AS website,
  array_to_string(websites, '|') AS websites_all,
  CASE WHEN length(phones) > 0 THEN phones[1] ELSE NULL END AS phone,
  array_to_string(phones, '|') AS phones_all,
  CASE WHEN length(emails) > 0 THEN emails[1] ELSE NULL END AS email,
  array_to_string(socials, '|') AS socials,
  brand, address_line, city, state, zip, country, lon, lat,
  source_dataset, source_id
FROM 'sonoma_places.parquet'
""")
con.execute("COPY flat TO 'sonoma_places.csv' (HEADER, DELIMITER ',')")
print(f"  -> wrote sonoma_places.csv")

# Excel summary — named businesses, chains excluded, tiered by lead quality.
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

print("Building Excel summary (chains excluded, tiered)...")
weak_sql = " OR ".join(f"lower(website) LIKE '%{w}%'" for w in WEAK_DOMAINS)
nonbiz_sql = ", ".join(f"'{c}'" for c in NON_BUSINESS_CATEGORIES)
# A name is "chain-like" if Overture flags a brand OR the same name recurs across
# the county at/above the threshold. Either way, skip it as a lead.
con.execute(f"""
CREATE TABLE leads AS
WITH name_counts AS (
  SELECT name, COUNT(*) AS n FROM flat WHERE name IS NOT NULL GROUP BY name
)
SELECT f.*,
  (f.brand IS NOT NULL AND f.brand <> '') OR nc.n >= {CHAIN_NAME_THRESHOLD} AS is_chain
FROM flat f
JOIN name_counts nc ON f.name = nc.name
WHERE f.name IS NOT NULL
""")
rows = con.execute(f"""
SELECT
  CASE
    WHEN website IS NULL OR website = '' THEN 'A'
    WHEN {weak_sql} THEN 'A'
    ELSE 'C'
  END AS tier,
  CASE
    WHEN website IS NULL OR website = '' THEN 'No website'
    WHEN {weak_sql} THEN 'Social/listing only — no owned site'
    ELSE 'Has a website — audit quality before pitching'
  END AS why,
  name, category_primary, city, phone, website, address_line,
  lon, lat, source_dataset, id
FROM leads
WHERE NOT is_chain                          -- chains/franchises and repeated-name POIs
  AND category_primary NOT IN ({nonbiz_sql})   -- parks, beaches, monuments, post offices…
ORDER BY tier ASC, name ASC
""").fetchall()

# How many chains/repeated POIs we dropped, for transparency
n_chains = con.execute("SELECT COUNT(*) FROM leads WHERE is_chain").fetchone()[0]
n_nonbiz = con.execute(
    f"SELECT COUNT(*) FROM leads WHERE NOT is_chain AND category_primary IN ({nonbiz_sql})"
).fetchone()[0]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sonoma Businesses"
headers = ["Tier", "Why", "Business", "Category", "City", "Phone", "Website",
           "Address", "Lon", "Lat", "Source", "Overture ID"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill("solid", start_color="1F4E78")
    c.alignment = Alignment(horizontal="center")
widths = [6, 34, 30, 26, 16, 16, 36, 40, 12, 12, 14, 28]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

TIER_A = PatternFill("solid", start_color="C6EFCE")
for r in rows:
    ws.append(r)
    if r[0] == "A":
        ws.cell(row=ws.max_row, column=1).fill = TIER_A
        ws.cell(row=ws.max_row, column=1).font = Font(name="Arial", bold=True)

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
wb.save("sonoma_businesses.xlsx")
n_a = sum(1 for r in rows if r[0] == "A")
print(f"  -> wrote sonoma_businesses.xlsx ({len(rows):,} businesses, {n_a:,} Tier-A; "
      f"excluded {n_chains:,} chains + {n_nonbiz:,} parks/landmarks/civic POIs)")

print("\n=========================================")
print("  DONE")
print("=========================================")
print(f"Files written to current folder:")
print(f"  sonoma_places.parquet  - raw Overture data ({n:,} records)")
print(f"  sonoma_places.csv      - flattened CSV")
print(f"  sonoma_businesses.xlsx - ranked summary ({len(rows):,} leads, chains removed)")
