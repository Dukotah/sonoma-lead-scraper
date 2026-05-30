"""
Enrich Sonoma leads: audit websites, web-verify "no website" leads, score, and export.

Reads the bulk Overture data produced by download_sonoma.py, removes chains and
non-business POIs, then for the leads you care about:
  - live-audits each existing website (HTTPS / mobile / load time / DIY builder)
  - web-searches each "no website" lead to confirm it really has no site
  - scores + tiers every lead and writes a ranked .xlsx + a CRM-ready .csv

You normally filter to the niche/city you're about to call, so you're not auditing
all 30k county records at once.

Examples
--------
  # Wineries with no real website, verify the no-site ones, export
  python enrich_leads.py --category winery --verify 30

  # Salons in Santa Rosa, audit their sites, cap at 100
  python enrich_leads.py --category salon --city "Santa Rosa" --limit 100

  # Combine with an OSM scrape exported from the desktop GUI
  python enrich_leads.py --category restaurant --merge-osm leads_santa_rosa.xlsx

Outputs: sonoma_leads_enriched.xlsx  and  sonoma_leads_crm.csv
"""
import sys, os, re, csv, time, argparse
from concurrent.futures import ThreadPoolExecutor, as_completed

# Windows console can't encode some characters under cp1252 — force UTF-8.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

try:
    import duckdb
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--quiet", "duckdb"])
    import duckdb

from lead_tools import audit_website, web_verify, score_lead, pitch_for, is_weak_url

# Same chain/non-business filters the downloader uses (kept here so this script
# can run standalone against the raw parquet).
CHAIN_NAME_THRESHOLD = 5
NON_BUSINESS_CATEGORIES = {
    "park", "beach", "river", "mountain", "mountain_peak", "hiking_trail", "trail",
    "lake", "forest", "island", "dam", "waterfall", "canyon", "valley", "cliff",
    "plateau", "natural_feature", "structure_and_geography", "body_of_water",
    "landmark_and_historical_building", "monument", "bridge", "cemetery",
    "post_office", "fire_station", "police_department",
}
MAX_AUDIT_WORKERS = 8


def norm(s: str) -> str:
    """Normalized key for dedupe across sources."""
    return re.sub(r"[^a-z0-9]+", "", (s or "").lower())


def load_overture(parquet: str, category=None, city=None, limit=None):
    con = duckdb.connect()
    nonbiz = ", ".join(f"'{c}'" for c in NON_BUSINESS_CATEGORIES)
    # download_sonoma.py writes already-projected columns (not nested Overture structs).
    con.execute(f"""
    CREATE TABLE flat AS
    SELECT id, name, category_primary AS category, brand,
      CASE WHEN length(websites) > 0 THEN websites[1] ELSE NULL END AS website,
      CASE WHEN length(phones)   > 0 THEN phones[1]   ELSE NULL END AS phone,
      address_line AS address, city,
      lon, lat
    FROM read_parquet('{parquet}')
    """)
    con.execute(f"""
    CREATE TABLE leads AS
    WITH nc AS (SELECT name, COUNT(*) n FROM flat WHERE name IS NOT NULL GROUP BY name)
    SELECT f.* FROM flat f JOIN nc ON f.name = nc.name
    WHERE f.name IS NOT NULL
      AND NOT ((f.brand IS NOT NULL AND f.brand <> '') OR nc.n >= {CHAIN_NAME_THRESHOLD})
      AND f.category NOT IN ({nonbiz})
    """)
    where, params = [], []
    if category:
        where.append("lower(category) LIKE ?")
        params.append(f"%{category.lower()}%")
    if city:
        where.append("lower(city) LIKE ?")
        params.append(f"%{city.lower()}%")
    sql = "SELECT name, category, website, phone, address, city, lon, lat FROM leads"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY (website IS NULL OR website = '') DESC, name ASC"
    if limit:
        sql += f" LIMIT {int(limit)}"
    cols = ["name", "category", "website", "phone", "address", "city", "lon", "lat"]
    return [dict(zip(cols, row)) for row in con.execute(sql, params).fetchall()]


def load_osm_export(path: str):
    """Read a GUI-exported leads xlsx/csv into the same dict shape for merging."""
    out = []
    if path.lower().endswith(".csv"):
        with open(path, newline="", encoding="utf-8") as f:
            for r in csv.DictReader(f):
                out.append({"name": r.get("Business") or r.get("name"),
                            "category": r.get("Niche") or r.get("category"),
                            "website": r.get("Website") or r.get("website") or "",
                            "phone": r.get("Phone") or r.get("phone") or "",
                            "address": r.get("Address") or "", "city": r.get("City") or "",
                            "lon": None, "lat": None})
        return out
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Leads"] if "Leads" in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip().lower() if h else "" for h in next(rows)]
    idx = {h: i for i, h in enumerate(header)}
    def g(row, *names):
        for n in names:
            if n in idx and row[idx[n]] is not None:
                return row[idx[n]]
        return ""
    for row in rows:
        nm = g(row, "business", "name")
        if not nm:
            continue
        out.append({"name": nm, "category": g(row, "niche", "category"),
                    "website": g(row, "website"), "phone": g(row, "phone"),
                    "address": g(row, "address"), "city": g(row, "city"),
                    "lon": None, "lat": None})
    return out


def main():
    ap = argparse.ArgumentParser(description="Enrich + verify + export Sonoma leads")
    ap.add_argument("--parquet", default="sonoma_places.parquet",
                    help="Overture file from download_sonoma.py")
    ap.add_argument("--category", help="filter Overture category (substring, e.g. 'winery')")
    ap.add_argument("--city", help="filter city (substring)")
    ap.add_argument("--limit", type=int, help="cap number of leads processed")
    ap.add_argument("--no-audit", action="store_true", help="skip live website audits")
    ap.add_argument("--verify", type=int, default=25,
                    help="web-verify up to N 'no website' leads (0 = skip)")
    ap.add_argument("--merge-osm", help="also merge a GUI-exported OSM leads .xlsx/.csv")
    ap.add_argument("--out", default="sonoma_leads", help="output filename stem")
    args = ap.parse_args()

    if not os.path.exists(args.parquet):
        sys.exit(f"ERROR: {args.parquet} not found. Run download_sonoma.py first.")

    print(f"Loading Overture leads from {args.parquet} ...")
    leads = load_overture(args.parquet, args.category, args.city, args.limit)
    for r in leads:
        r["source"] = "overture"
    print(f"  {len(leads)} leads after chain/non-business filtering"
          + (f" (category~{args.category})" if args.category else "")
          + (f" (city~{args.city})" if args.city else ""))

    # Combine with an OSM export, deduped by normalized name + city
    if args.merge_osm:
        if not os.path.exists(args.merge_osm):
            sys.exit(f"ERROR: --merge-osm file {args.merge_osm} not found.")
        osm = load_osm_export(args.merge_osm)
        # Dedupe on normalized name alone: within one county the same business often
        # carries different/missing city values across Overture vs OSM, so name+city
        # would miss real duplicates. Same normalized name in a county ≈ same business.
        seen = {norm(r["name"]) for r in leads}
        added = 0
        for r in osm:
            key = norm(r["name"])
            if not key or key in seen:
                continue
            seen.add(key)
            r["source"] = "osm"
            leads.append(r)
            added += 1
        print(f"  merged OSM export: +{added} new businesses (deduped against Overture)")

    # Live-audit existing websites (parallel, capped)
    audits = {}
    if not args.no_audit:
        sites = [(i, r["website"]) for i, r in enumerate(leads)
                 if r.get("website") and not is_weak_url(r["website"])[0]]
        print(f"Auditing {len(sites)} real websites (parallel)...")
        with ThreadPoolExecutor(max_workers=MAX_AUDIT_WORKERS) as ex:
            futs = {ex.submit(audit_website, w): i for i, w in sites}
            done = 0
            for f in as_completed(futs):
                audits[futs[f]] = f.result()
                done += 1
                if done % 25 == 0:
                    print(f"  audited {done}/{len(sites)}")

    # Web-verify a slice of the "no website" leads
    if args.verify:
        nosite = [r for r in leads if not r.get("website")][:args.verify]
        print(f"Web-verifying {len(nosite)} 'no website' leads (~2s each)...")
        n_false, n_throttled = 0, 0
        for i, r in enumerate(nosite, 1):
            v = web_verify(r["name"], r.get("city", ""))
            r["verify"] = v
            if v["verdict"] == "has_site":
                r["verify_note"] = f"FALSE POSITIVE? found {v['likely_site']}"
                n_false += 1
            elif v["verdict"] == "no_site_found":
                r["verify_note"] = "confirmed: no owned site in search results"
            elif v["verdict"] == "throttled":
                r["verify_note"] = "search throttled — verify manually"
                n_throttled += 1
            else:
                r["verify_note"] = "could not verify"
            time.sleep(2.0)  # pace requests; the search endpoint throttles bursts
        print(f"  verified: {n_false} false-positives (have a site), "
              f"{n_throttled} throttled/unverified")
        if n_throttled > len(nosite) // 2:
            print("  NOTE: search is rate-limiting this IP. Re-run later or use a "
                  "smaller --verify batch; results above are still valid.")

    # Score + tier + pitch
    for i, r in enumerate(leads):
        s, why, tier = score_lead(r, audits.get(i, {}))
        # a confirmed false-positive (site found in search) shouldn't sit in Tier A
        if r.get("verify", {}).get("verdict") == "has_site":
            tier = "C"
            why += "; site found via web search"
        r["score"], r["score_reasons"], r["tier"] = s, why, tier
        r["pitch"] = pitch_for(r, tier)
    leads.sort(key=lambda x: -x["score"])

    n = {"A": 0, "B": 0, "C": 0}
    for r in leads:
        n[r["tier"]] = n.get(r["tier"], 0) + 1
    print(f"Scored {len(leads)} leads — Tier A={n['A']}  B={n['B']}  C={n['C']}")

    write_outputs(leads, audits, args.out)


def write_outputs(leads, audits, stem):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # CRM CSV (HubSpot/Pipedrive-friendly column names)
    csv_path = f"{stem}_crm.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Company name", "Phone number", "Website URL", "City", "Address",
                    "Lead tier", "Lead score", "Source", "Why a lead",
                    "Verification", "Pitch"])
        for r in leads:
            w.writerow([r.get("name"), r.get("phone"), r.get("website"), r.get("city"),
                        r.get("address"), r.get("tier"), r.get("score"), r.get("source"),
                        r.get("score_reasons"), r.get("verify_note", ""), r.get("pitch")])
    print(f"  -> wrote {csv_path}")

    # Pretty xlsx
    wb = openpyxl.Workbook()
    s = wb.active
    s.title = "Leads"
    HEAD_F = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    HEAD_FILL = PatternFill("solid", start_color="1F4E78")
    BODY = Font(name="Arial", size=10)
    WRAP = Alignment(wrap_text=True, vertical="top")
    TIER_FILL = {"A": PatternFill("solid", start_color="C6EFCE"),
                 "B": PatternFill("solid", start_color="FFEB9C"),
                 "C": PatternFill("solid", start_color="FFC7CE")}
    cols = [("Tier", "tier", 7), ("Score", "score", 7), ("Business", "name", 30),
            ("Category", "category", 20), ("City", "city", 16), ("Phone", "phone", 16),
            ("Website", "website", 34), ("Source", "source", 10),
            ("Why a lead", "score_reasons", 34), ("Verification", "verify_note", 34),
            ("Pitch", "pitch", 55), ("Address", "address", 32)]
    for i, (h, _, w) in enumerate(cols, 1):
        c = s.cell(row=1, column=i, value=h)
        c.font = HEAD_F; c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        s.column_dimensions[get_column_letter(i)].width = w
    for ri, r in enumerate(leads, 2):
        for i, (_, k, _) in enumerate(cols, 1):
            c = s.cell(row=ri, column=i, value=r.get(k))
            c.font = BODY; c.alignment = WRAP
        if r["tier"] in TIER_FILL:
            s.cell(row=ri, column=1).fill = TIER_FILL[r["tier"]]
    s.freeze_panes = "A2"
    s.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(leads)+1}"
    xlsx_path = f"{stem}_enriched.xlsx"
    wb.save(xlsx_path)
    print(f"  -> wrote {xlsx_path}")


if __name__ == "__main__":
    main()
