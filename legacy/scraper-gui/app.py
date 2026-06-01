"""
Lead Scraper — finds local businesses that need websites.
Source: OpenStreetMap (Overpass API) — free, unlimited, never breaks.
"""
import os
import re
import io
import json
import time
import socket
import threading
from collections import OrderedDict, Counter
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse

import requests
from flask import Flask, render_template_string, request, send_file, jsonify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)

UA = "LeadScraper/1.0 (contact: dukotah@gmail.com)"
AUDIT_TIMEOUT = 6
MAX_AUDIT_WORKERS = 8

# Each niche -> list of OSM tag filters.
NICHE_OPTIONS = {
    "Home services": [
        ("plumbers",          "Plumbers",          ['craft=plumber']),
        ("electricians",      "Electricians",      ['craft=electrician']),
        ("hvac",              "HVAC",              ['craft=hvac', 'shop=hvac']),
        ("roofers",           "Roofers",           ['craft=roofer']),
        ("painters",          "Painters",          ['craft=painter']),
        ("carpenters",        "Carpenters",        ['craft=carpenter']),
        ("landscapers",       "Landscapers",       ['craft=gardener', 'shop=garden_centre']),
        ("pest_control",      "Pest Control",      ['craft=pest_control']),
        ("locksmith",         "Locksmiths",        ['shop=locksmith']),
    ],
    "Auto": [
        ("car_repair",        "Auto Repair",       ['shop=car_repair']),
        ("car_parts",         "Auto Parts",        ['shop=car_parts']),
        ("tire_shops",        "Tire Shops",        ['shop=tyres']),
        ("car_wash",          "Car Wash",          ['amenity=car_wash']),
        ("motorcycle_repair", "Motorcycle Repair", ['shop=motorcycle_repair']),
    ],
    "Personal services": [
        ("hairdresser",       "Hair Salons",       ['shop=hairdresser']),
        ("beauty",            "Beauty Salons",     ['shop=beauty']),
        ("massage",           "Massage",           ['shop=massage', 'amenity=spa']),
        ("dry_cleaning",      "Dry Cleaners",      ['shop=dry_cleaning', 'shop=laundry']),
        ("tattoo",            "Tattoo Shops",      ['shop=tattoo']),
    ],
    "Food": [
        ("restaurant",        "Restaurants",       ['amenity=restaurant']),
        ("cafe",              "Cafes / Coffee",    ['amenity=cafe']),
        ("bakery",            "Bakeries",          ['shop=bakery']),
        ("fast_food",         "Fast Food",         ['amenity=fast_food']),
        ("bar",               "Bars / Pubs",       ['amenity=bar', 'amenity=pub']),
        ("ice_cream",         "Ice Cream",         ['amenity=ice_cream']),
        ("butcher",           "Butcher",           ['shop=butcher']),
        ("deli",              "Deli",              ['shop=deli']),
    ],
    "Specialty retail": [
        ("florist",           "Florists",          ['shop=florist']),
        ("jewelry",           "Jewelry",           ['shop=jewelry']),
        ("antiques",          "Antiques",          ['shop=antiques']),
        ("books",             "Books",             ['shop=books']),
        ("clothes",           "Clothing",          ['shop=clothes']),
        ("furniture",         "Furniture",         ['shop=furniture']),
        ("gift",              "Gift Shops",        ['shop=gift']),
        ("hardware",          "Hardware",          ['shop=hardware']),
        ("bicycle",           "Bicycle Shops",     ['shop=bicycle']),
        ("pet",               "Pet Supplies",      ['shop=pet']),
    ],
    "Professional / medical": [
        ("dentist",           "Dentists",          ['amenity=dentist', 'healthcare=dentist']),
        ("doctors",           "Doctors",           ['amenity=doctors', 'healthcare=doctor']),
        ("veterinary",        "Veterinarians",     ['amenity=veterinary']),
        ("lawyer",            "Attorneys",         ['office=lawyer']),
        ("accountant",        "Accountants",       ['office=accountant']),
        ("real_estate",       "Real Estate",       ['office=estate_agent']),
        ("insurance",         "Insurance",         ['office=insurance']),
        ("photographer",      "Photographers",     ['craft=photographer', 'shop=photo']),
    ],
}

# ============ OVERPASS / NOMINATIM ============
NOMINATIM = "https://nominatim.openstreetmap.org/search"
OVERPASS_ENDPOINTS = [
    "https://overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
    "https://overpass.private.coffee/api/interpreter",
]

def geocode_city(city_str: str) -> dict | None:
    """Returns {lat, lon, bbox=[south, west, north, east], display_name}."""
    try:
        r = requests.get(NOMINATIM, params={"q": city_str, "format": "json", "limit": 1},
                         headers={"User-Agent": UA}, timeout=10)
        if r.status_code != 200 or not r.json():
            return None
        d = r.json()[0]
        # nominatim bbox = [south, north, west, east]
        bb = [float(x) for x in d["boundingbox"]]
        return {
            "lat": float(d["lat"]), "lon": float(d["lon"]),
            "bbox": (bb[0], bb[2], bb[1], bb[3]),  # south, west, north, east
            "display_name": d["display_name"],
        }
    except Exception:
        return None


# Transient HTTP statuses worth retrying on the SAME endpoint before moving on.
# (429 = rate limited, 502/503/504 = mirror overloaded — all usually clear up.)
RETRY_STATUS = {429, 502, 503, 504}


def overpass_query(bbox: tuple, tag_filters: list, timeout: int = 60) -> list:
    """Run an Overpass query for nwr matching any tag in tag_filters within bbox.

    Rotates through OVERPASS_ENDPOINTS. Transient failures (rate-limit / gateway
    timeout / read timeout) are retried once on the same mirror with a short
    backoff; hard rejections (403/406) skip straight to the next mirror. Raises
    RuntimeError with a per-mirror breakdown only if every mirror fails.
    """
    south, west, north, east = bbox
    parts = []
    for tf in tag_filters:
        if "=" not in tf:
            continue
        k, v = tf.split("=", 1)
        parts.append(f'nwr["{k}"="{v}"]({south},{west},{north},{east});')
    if not parts:
        return []
    body = f"[out:json][timeout:{timeout}];\n(\n  " + "\n  ".join(parts) + "\n);\nout center tags;"
    errors = []
    for endpoint in OVERPASS_ENDPOINTS:
        host = _hostname(endpoint) or endpoint
        for attempt in range(2):  # initial try + one retry for transient errors
            try:
                r = requests.post(endpoint, data=body, headers={"User-Agent": UA},
                                  timeout=timeout + 10)
                if r.status_code == 200:
                    return r.json().get("elements", [])
                errors.append(f"{host} HTTP {r.status_code}")
                if r.status_code in RETRY_STATUS and attempt == 0:
                    time.sleep(2)
                    continue  # retry same mirror
                break  # hard rejection — move to next mirror
            except requests.exceptions.RequestException as e:
                errors.append(f"{host} {type(e).__name__}")
                if attempt == 0:
                    time.sleep(2)
                    continue  # transient network error — retry once
                break
    raise RuntimeError(
        "All Overpass mirrors failed (" + "; ".join(errors) + "). "
        "This is usually temporary mirror overload — wait a minute and retry. "
        "If it persists on a cloud/VPN IP, some mirrors block those; try a home connection."
    )


def osm_to_lead(el: dict, niche_label: str) -> dict | None:
    tags = el.get("tags", {})
    name = tags.get("name") or tags.get("operator")
    if not name:
        return None  # unnamed amenities aren't useful leads
    # Skip national chains/franchises — a `brand` tag means corporate marketing
    # owns the website decision; they won't buy from a local web designer.
    if tags.get("brand") or tags.get("brand:wikidata"):
        return None
    addr_parts = []
    for k in ("addr:housenumber", "addr:street"):
        if tags.get(k):
            addr_parts.append(tags[k])
    line1 = " ".join(addr_parts).strip()
    city = tags.get("addr:city", "")
    state = tags.get("addr:state", "")
    postcode = tags.get("addr:postcode", "")
    full_addr = ", ".join(p for p in [line1, city, f"{state} {postcode}".strip()] if p)
    osm_id = f"{el['type'][0]}{el['id']}"
    return {
        "name": name.strip(),
        "slug": osm_id,
        "website": (tags.get("website") or tags.get("contact:website") or "").strip(),
        "phone": (tags.get("phone") or tags.get("contact:phone") or "").strip(),
        "address": full_addr,
        "city": city,
        "years_in_business": "",
        "niche": niche_label,
        "osm_url": f"https://www.openstreetmap.org/{el['type']}/{el['id']}",
        "lat": el.get("lat") or (el.get("center") or {}).get("lat"),
        "lon": el.get("lon") or (el.get("center") or {}).get("lon"),
    }


def osm_scrape(city_geo: dict, niche_keys: list, progress=None) -> list:
    """Query Overpass per-niche to keep requests small and responsive."""
    all_niches = {k: (label, tags) for grp in NICHE_OPTIONS.values()
                  for (k, label, tags) in grp}
    leads = []
    for i, key in enumerate(niche_keys, 1):
        if key not in all_niches:
            continue
        label, tags = all_niches[key]
        if progress:
            progress(i, len(niche_keys), f"Querying OSM for {label}…")
        try:
            els = overpass_query(city_geo["bbox"], tags)
            n_named = 0
            for el in els:
                rec = osm_to_lead(el, label)
                if rec:
                    leads.append(rec)
                    n_named += 1
            if progress:
                progress(i, len(niche_keys), f"  {label}: {n_named} named businesses")
        except Exception as e:
            if progress:
                progress(i, len(niche_keys), f"  {label}: ERROR {e}")
        time.sleep(1.0)  # be polite to Overpass
    return leads


# ============ WEBSITE AUDIT ============
WEAK_DOMAINS = (
    "yelp.com", "yelp.to", "facebook.com", "fb.com", "instagram.com",
    "yellowpages.com", "localsearch.com", "google.com", "youtube.com",
    "linkedin.com", "twitter.com", "x.com", "tiktok.com", "pinterest.com",
    "groupon.com", "locality.com", "nextdoor.com",
)

def _hostname(url: str) -> str:
    """Lowercased hostname for a URL, tolerant of missing scheme."""
    if not url:
        return ""
    u = url if "://" in url else "http://" + url
    try:
        return (urlparse(u).hostname or "").lower()
    except Exception:
        return ""


def is_weak_url(url: str):
    if not url:
        return True, "no website"
    host = _hostname(url)
    # match on domain boundary, not substring — "x.com" must not flag "fedex.com"
    for d in WEAK_DOMAINS:
        if host == d or host.endswith("." + d):
            return True, d
    return False, ""


def audit_website(url: str) -> dict:
    out = {"reachable": False, "https": False, "load_ms": None,
           "mobile_viewport": False, "builder": "", "title": "",
           "size_kb": None, "audit_notes": []}
    if not url:
        return out
    if not url.startswith("http"):
        url = "http://" + url
    out["https"] = url.startswith("https://")
    try:
        t0 = time.time()
        r = requests.get(url, headers={"User-Agent": UA},
                         timeout=AUDIT_TIMEOUT, allow_redirects=True)
        ms = int((time.time() - t0) * 1000)
        out["load_ms"] = ms
        out["reachable"] = r.status_code < 400
        out["https"] = r.url.startswith("https://")
        out["size_kb"] = round(len(r.content) / 1024, 1)
        html = r.text[:50000].lower()
        if 'name="viewport"' in html:
            out["mobile_viewport"] = True
        tm = re.search(r"<title>([^<]+)</title>", r.text, re.IGNORECASE)
        if tm:
            out["title"] = tm.group(1).strip()[:100]
        for sig, b in [("wix.com","Wix"),("squarespace","Squarespace"),("weebly","Weebly"),
                       ("godaddy","GoDaddy Sites"),("wordpress","WordPress"),
                       ("shopify","Shopify"),("webflow","Webflow"),("duda","Duda"),
                       ("site123","Site123"),("jimdo","Jimdo")]:
            if sig in html:
                out["builder"] = b; break
        if ms > 4000: out["audit_notes"].append(f"Slow load ({ms}ms)")
        if not out["mobile_viewport"]: out["audit_notes"].append("No mobile viewport")
        if not out["https"]: out["audit_notes"].append("No HTTPS")
        if out["builder"] in ("Wix","Weebly","GoDaddy Sites","Site123","Jimdo"):
            out["audit_notes"].append(f"DIY-builder ({out['builder']})")
    except requests.exceptions.SSLError:
        out["audit_notes"].append("SSL / broken cert")
    except requests.exceptions.Timeout:
        out["audit_notes"].append(f"Timeout (>{AUDIT_TIMEOUT}s)")
    except Exception as e:
        out["audit_notes"].append(f"Unreachable: {type(e).__name__}")
    return out


def score_lead(rec, audit):
    score, reasons = 0, []
    is_weak, why = is_weak_url(rec["website"])
    if not rec["website"]:
        score += 60; reasons.append("NO WEBSITE on OSM")
        tier = "A"
    elif is_weak:
        score += 40; reasons.append(f"non-site link ({why})")
        tier = "A"
    else:
        tier = "C"
        if not audit.get("reachable"):
            score += 50; reasons.append("Site listed but unreachable"); tier = "A"
        else:
            if not audit.get("https"):
                score += 18; reasons.append("HTTP only (no SSL)"); tier = "B"
            if not audit.get("mobile_viewport"):
                score += 14; reasons.append("Not mobile-friendly"); tier = "B"
            if audit.get("load_ms") and audit["load_ms"] > 4000:
                score += 10; reasons.append(f"Slow ({audit['load_ms']}ms)"); tier = "B"
            if audit.get("builder") in ("Wix","Weebly","GoDaddy Sites","Site123","Jimdo"):
                score += 12; reasons.append(f"DIY-builder ({audit['builder']})"); tier = "B"
            if not reasons:
                reasons.append("Real site, no obvious issues")
    if rec.get("phone"): score += 4; reasons.append("phone listed")
    return score, "; ".join(reasons), tier


def pitch_for(rec, score, tier):
    if not rec["website"]:
        return f"No website on OSM. Verify by Googling '{rec['name']} {rec['city']}' — if nothing comes up, pitch: 1-page site that ranks for '{rec['niche'].lower()} {rec['city']}' + click-to-call."
    u = rec["website"].lower()
    if "yelp" in u: return "Yelp-only. Pitch: own your domain, stop paying Yelp ad fees."
    if "facebook" in u: return "Facebook-only. Pitch: real site = Google rankings, social complements."
    if u.startswith("http://"): return "HTTP only — Chrome warns visitors. Quick rebuild + SSL."
    return "Has real site — verify quality manually before pitching."


# ============ XLSX EXPORT ============
def build_xlsx(leads, audits, params):
    wb = Workbook()
    HEAD_F = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    HEAD_FILL = PatternFill("solid", start_color="1F4E78")
    BODY = Font(name="Arial", size=10)
    WRAP = Alignment(wrap_text=True, vertical="top")
    CENTER = Alignment(horizontal="center", vertical="top")
    TIER_FILL = {"A": PatternFill("solid", start_color="C6EFCE"),
                 "B": PatternFill("solid", start_color="FFEB9C"),
                 "C": PatternFill("solid", start_color="FFC7CE")}
    s = wb.active; s.title = "Leads"
    cols = [
        ("Tier","tier",7),("Score","score",7),("Business","name",30),
        ("Niche","niche",18),("City","city",16),("Phone","phone",16),
        ("Website","website",36),("Reachable","ar",10),("HTTPS","ah",8),
        ("Mobile","am",9),("Load (ms)","al",10),("Builder","ab",13),
        ("Audit notes","an",32),("Why a lead","score_reasons",36),
        ("Pitch","pitch",55),("Address","address",34),
        ("OSM link","osm_url",30),
    ]
    for i,(h,_,w) in enumerate(cols,1):
        c = s.cell(row=1,column=i,value=h)
        c.font = HEAD_F; c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        s.column_dimensions[get_column_letter(i)].width = w
    for ri,rec in enumerate(leads,2):
        a = audits.get(rec["slug"], {})
        row = {
            "tier":rec.get("tier"),"score":rec.get("score"),
            "name":rec["name"],"niche":rec["niche"],"city":rec.get("city"),
            "phone":rec.get("phone"),"website":rec.get("website"),
            "ar":"Yes" if a.get("reachable") else ("No" if rec.get("website") else ""),
            "ah":"Yes" if a.get("https") else ("No" if rec.get("website") else ""),
            "am":"Yes" if a.get("mobile_viewport") else ("No" if rec.get("website") else ""),
            "al":a.get("load_ms"),"ab":a.get("builder",""),
            "an":"; ".join(a.get("audit_notes",[])),
            "score_reasons":rec.get("score_reasons"),"pitch":rec.get("pitch"),
            "address":rec.get("address"),"osm_url":rec.get("osm_url"),
        }
        for i,(_,k,_) in enumerate(cols,1):
            c = s.cell(row=ri,column=i,value=row.get(k))
            c.font = BODY; c.alignment = WRAP
        tier = rec.get("tier","C")
        if tier in TIER_FILL:
            s.cell(row=ri,column=1).fill = TIER_FILL[tier]
            s.cell(row=ri,column=1).alignment = CENTER
            s.cell(row=ri,column=1).font = Font(name="Arial",bold=True,size=10)
    s.freeze_panes = "A2"
    s.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(leads)+1}"

    sm = wb.create_sheet("Summary")
    sm["A1"] = f"OSM scrape — {params.get('city','')} — {datetime.now():%Y-%m-%d %H:%M}"
    sm["A1"].font = Font(name="Arial",bold=True,size=14)
    sm["A3"]="Total leads"; sm["B3"]=len(leads)
    sm["A4"]="Tier A"; sm["B4"]=f'=COUNTIF(Leads!A2:A{len(leads)+1},"A")'
    sm["A5"]="Tier B"; sm["B5"]=f'=COUNTIF(Leads!A2:A{len(leads)+1},"B")'
    sm["A6"]="Tier C"; sm["B6"]=f'=COUNTIF(Leads!A2:A{len(leads)+1},"C")'
    sm["A8"]="Source"; sm["B8"]="OpenStreetMap (Overpass API)"
    sm["A9"]="Niches"; sm["B9"]=", ".join(params.get("niches",[]))
    sm.column_dimensions["A"].width=20; sm.column_dimensions["B"].width=80
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()


# ============ FLASK / UI ============
JOBS = {}
INDEX_HTML = """<!doctype html>
<html><head><title>Lead Scraper</title>
<style>
body{font-family:-apple-system,BlinkMacSystemFont,Arial,sans-serif;max-width:920px;margin:24px auto;padding:0 16px;color:#222}
h1{margin-bottom:4px}.sub{color:#666;margin-bottom:16px}
.note{background:#fff8e1;border-left:4px solid #f0c020;padding:10px 14px;margin:12px 0;border-radius:4px;font-size:13px}
fieldset{border:1px solid #ddd;border-radius:8px;margin:14px 0;padding:14px}
legend{padding:0 8px;font-weight:600}
label{display:inline-block;margin:3px 10px 3px 0;font-size:14px;cursor:pointer}
input[type=text]{padding:6px 8px;font-size:14px;border:1px solid #bbb;border-radius:4px;width:300px}
button{background:#1F4E78;color:#fff;border:none;padding:10px 18px;border-radius:6px;font-size:15px;cursor:pointer}
button:hover{background:#143b5d}button:disabled{background:#aaa;cursor:not-allowed}
.row{margin:8px 0}.hint{color:#888;font-size:12px}
#status{margin-top:18px;padding:12px;background:#f5f5f5;border-radius:6px;display:none;
        font-family:Menlo,Consolas,monospace;font-size:12px;max-height:260px;overflow-y:auto;white-space:pre}
#download{display:none;margin-top:14px}
a.dl{display:inline-block;padding:10px 18px;background:#28a745;color:#fff;text-decoration:none;border-radius:6px}
</style></head><body>
<h1>Lead Scraper</h1>
<div class="sub">Find local businesses that need websites. Source: OpenStreetMap.</div>
<div class="note"><b>About OSM data:</b> coverage depends on local volunteers, so you'll get fewer raw results than Yellow Pages
— but the data is real, never breaks, and businesses with no website tag are real "no website" candidates worth verifying.</div>

<form id="form">
  <fieldset>
    <legend>Where</legend>
    <div class="row">
      City: <input type="text" id="city" value="Santa Rosa, California" placeholder="Santa Rosa, California">
      <span class="hint">Real city name. Include state/country for clarity.</span>
    </div>
  </fieldset>

  <fieldset>
    <legend>Niches — pick any</legend>
    {{ niche_html|safe }}
  </fieldset>

  <fieldset>
    <legend>Website audit</legend>
    <label><input type="checkbox" id="audit" checked> Live-fetch each lead's website (slower, real quality scores)</label>
  </fieldset>

  <button type="submit" id="go">Scrape</button>
</form>

<div id="status"></div>
<div id="download"></div>

<script>
const form=document.getElementById("form"),go=document.getElementById("go"),status=document.getElementById("status"),dl=document.getElementById("download");
form.addEventListener("submit", async (e)=>{
  e.preventDefault();
  const city=document.getElementById("city").value.trim();
  const audit=document.getElementById("audit").checked;
  const niches=[...document.querySelectorAll("input[name=niche]:checked")].map(x=>x.value);
  if(niches.length===0){alert("Pick at least one niche.");return;}
  go.disabled=true;go.textContent="Scraping…";
  status.style.display="block";status.textContent="Starting…";dl.style.display="none";
  const res=await fetch("/scrape",{method:"POST",headers:{"Content-Type":"application/json"},
    body:JSON.stringify({city,audit,niches})});
  const {job_id}=await res.json();
  const poll=setInterval(async ()=>{
    const r=await fetch(`/progress/${job_id}`);const j=await r.json();
    status.textContent=j.log.join("\\n");status.scrollTop=status.scrollHeight;
    if(j.done){clearInterval(poll);go.disabled=false;go.textContent="Scrape";
      if(j.file){dl.style.display="block";
        dl.innerHTML=`<a class="dl" href="/download/${job_id}">Download ${j.file}</a>`;}}
  },800);
});
</script>
</body></html>"""

def render_niches():
    parts=[]
    for group, items in NICHE_OPTIONS.items():
        parts.append(f"<div class='row'><strong>{group}:</strong><br>")
        for key, label, _ in items:
            parts.append(f'<label><input type="checkbox" name="niche" value="{key}"> {label}</label>')
        parts.append("</div>")
    return "\n".join(parts)


@app.route("/")
def index():
    return render_template_string(INDEX_HTML, niche_html=render_niches())


def run_job(job_id, city, audit_flag, niches):
    job = JOBS[job_id]
    def log(msg): job["log"].append(f"[{datetime.now():%H:%M:%S}] {msg}")

    log(f"Geocoding '{city}'…")
    geo = geocode_city(city)
    if not geo:
        log(f"ERROR: could not geocode '{city}'. Try including the state/country, e.g. 'Santa Rosa, California'.")
        job["done"] = True; return
    log(f"  → {geo['display_name']}")
    log(f"  bbox: S={geo['bbox'][0]:.3f} W={geo['bbox'][1]:.3f} N={geo['bbox'][2]:.3f} E={geo['bbox'][3]:.3f}")

    def progress(done, total, msg): log(f"({done}/{total}) {msg}")
    leads = osm_scrape(geo, niches, progress=progress)
    log(f"Total raw OSM listings: {len(leads)}")

    seen = OrderedDict()
    for r in leads:
        k = r["slug"]
        if k in seen:
            p = seen[k]
            if r["niche"] not in p["niche"]:
                p["niche"] = p["niche"] + " / " + r["niche"]
        else:
            seen[k] = r
    leads = list(seen.values())
    log(f"After dedup: {len(leads)}")

    audits = {}
    if audit_flag:
        sites = [(r["slug"], r["website"]) for r in leads if r.get("website")]
        log(f"Auditing {len(sites)} websites (parallel)…")
        with ThreadPoolExecutor(max_workers=MAX_AUDIT_WORKERS) as ex:
            futs = {ex.submit(audit_website, w): s for s, w in sites}
            done = 0
            for f in as_completed(futs):
                slug = futs[f]
                try: audits[slug] = f.result()
                except Exception: audits[slug] = {"audit_notes":["audit failed"]}
                done += 1
                if done % 10 == 0:
                    log(f"  audited {done}/{len(sites)}")
        log("Audits done.")
    else:
        log("Skipping website audits.")

    log("Scoring leads…")
    for r in leads:
        s, why, tier = score_lead(r, audits.get(r["slug"], {}))
        r["score"] = s; r["score_reasons"] = why; r["tier"] = tier
        r["pitch"] = pitch_for(r, s, tier)
    leads.sort(key=lambda x: -x["score"])
    log(f"Qualified leads: {len(leads)} "
        f"(A={sum(1 for r in leads if r['tier']=='A')}, "
        f"B={sum(1 for r in leads if r['tier']=='B')}, "
        f"C={sum(1 for r in leads if r['tier']=='C')})")

    log("Building xlsx…")
    fn = f"leads_{re.sub(r'[^a-z0-9]+', '_', city.lower())}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    job["file_bytes"] = build_xlsx(leads, audits, {"city": city, "niches": niches})
    job["filename"] = fn
    log(f"Done. → {fn}")
    job["done"] = True


@app.route("/scrape", methods=["POST"])
def scrape():
    data = request.get_json()
    jid = str(int(time.time()*1000))
    JOBS[jid] = {"log": [], "done": False, "file_bytes": None, "filename": None}
    threading.Thread(target=run_job, daemon=True,
                     args=(jid, data["city"], data["audit"], data["niches"])).start()
    return jsonify({"job_id": jid})


@app.route("/progress/<jid>")
def progress(jid):
    j = JOBS.get(jid)
    if not j: return jsonify({"error":"not found"}),404
    return jsonify({"log": j["log"], "done": j["done"], "file": j["filename"]})


@app.route("/download/<jid>")
def download(jid):
    j = JOBS.get(jid)
    if not j or not j.get("file_bytes"): return "Not ready", 404
    return send_file(io.BytesIO(j["file_bytes"]), download_name=j["filename"],
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def free_port(default=5000):
    for p in (default, 5001, 5050, 8000, 8080):
        with socket.socket() as s:
            try: s.bind(("127.0.0.1", p)); return p
            except OSError: continue
    return default


if __name__ == "__main__":
    port = free_port()
    print(f"\n=========================================")
    print(f"  Lead Scraper (OSM) — http://localhost:{port}")
    print(f"=========================================\n")
    app.run(host="127.0.0.1", port=port, debug=False)
