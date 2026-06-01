"""
Export — write scored leads to a CRM-ready CSV and a color-tiered XLSX.
Columns are defined by the vertical (vertical.columns), so the same exporter
serves every use case. Generalized from data-kit/enrich_leads.py::write_outputs.
"""
from __future__ import annotations

import csv


def write_csv(leads: list[dict], columns: list[tuple[str, str]], path: str) -> str:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([h for h, _ in columns])
        for r in leads:
            w.writerow([r.get(k, "") for _, k in columns])
    return path


def write_xlsx(leads: list[dict], columns: list[tuple[str, str]], path: str,
               title: str = "Leads") -> str:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    s = wb.active
    s.title = title[:31]
    HEAD_F = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    HEAD_FILL = PatternFill("solid", start_color="1F4E78")
    BODY = Font(name="Arial", size=10)
    WRAP = Alignment(wrap_text=True, vertical="top")
    TIER_FILL = {"A": PatternFill("solid", start_color="C6EFCE"),
                 "B": PatternFill("solid", start_color="FFEB9C"),
                 "C": PatternFill("solid", start_color="FFC7CE")}
    for i, (h, _) in enumerate(columns, 1):
        c = s.cell(row=1, column=i, value=h)
        c.font = HEAD_F; c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        s.column_dimensions[get_column_letter(i)].width = max(10, min(48, len(h) + 8))
    tier_col = next((i for i, (_, k) in enumerate(columns, 1) if k == "tier"), None)
    for ri, r in enumerate(leads, 2):
        for i, (_, k) in enumerate(columns, 1):
            c = s.cell(row=ri, column=i, value=r.get(k))
            c.font = BODY; c.alignment = WRAP
        if tier_col and r.get("tier") in TIER_FILL:
            s.cell(row=ri, column=tier_col).fill = TIER_FILL[r["tier"]]
    s.freeze_panes = "A2"
    if leads:
        s.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(leads)+1}"
    wb.save(path)
    return path


def write_outputs(leads, columns, stem: str, log=print) -> tuple[str, str]:
    csv_path = write_csv(leads, columns, f"{stem}_crm.csv")
    log(f"  -> {csv_path}")
    xlsx_path = write_xlsx(leads, columns, f"{stem}.xlsx")
    log(f"  -> {xlsx_path}")
    return csv_path, xlsx_path
