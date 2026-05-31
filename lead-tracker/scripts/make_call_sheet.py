"""
Build a Monday-morning call sheet of the WARMEST, callable web-design leads.

Ranks leads by an `outreach_score` that combines:
  - website NEED      : Tier A (no/social-only site) > Tier B (DIY builder) > C
  - REACHABILITY      : has phone (required) + email + socials
  - INDUSTRY FIT      : how likely the niche is to buy a site AND depend on being
                        found online (salon/restaurant/trades/medical = high;
                        professional services = medium; church/gov/nonprofit = low)
  - COMPLETENESS      : how filled-in the record is, + Overture confidence

It writes the score back into leads (outreach_score, industry_fit) so the tracker
can sort by it, and produces ready-to-use deliverables in data/:
  Monday_Call_Sheet.xlsx  - top callable leads, ranked, with pitch + blank
                            "Call outcome"/"Notes" columns; tabs: Top Calls,
                            Tier-B Upsell, By Niche
  warm_leads.csv          - every reachable warm lead, ranked

Run:  python scripts/make_call_sheet.py            (default: top 300 in xlsx)
      python scripts/make_call_sheet.py --top 500
"""
import os, re, sqlite3, argparse

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.normpath(os.path.join(HERE, "..", "data"))
DB_PATH = os.environ.get("LEADS_DB", os.path.join(DATA, "leads.sqlite"))

# Industry fit by keyword. A category matches the first bucket whose keywords
# appear in it. There are ~1,100 distinct Overture categories, so we match on
# substrings rather than enumerate them all.
HIGH_FIT = [
    "restaurant", "cafe", "coffee", "bar", "pub", "brewery", "winery", "wine",
    "bakery", "food", "caterer", "catering", "pizz", "taqueria", "deli", "diner",
    "salon", "barber", "hair", "nail", "spa", "beauty", "massage", "skin_care",
    "tattoo", "esthetic", "lash", "wax", "tanning",
    "dentist", "dental", "chiropract", "optometr", "veterinar", "physical_therapy",
    "dermatolog", "acupunctur", "orthodont", "medical_spa", "cosmetic",
    "plumb", "electric", "hvac", "roofing", "contractor", "landscap", "construction",
    "remodel", "painter", "flooring", "fencing", "paving", "concrete", "handyman",
    "pest_control", "locksmith", "excavat", "well_drilling", "septic", "solar",
    "auto_repair", "automotive", "mechanic", "tire", "body_shop", "detailing",
    "real_estate_agent", "realtor", "mortgage", "insurance_agent",
    "photograph", "videograph", "florist", "flowers", "event", "wedding", "dj_",
    "gym", "fitness", "yoga", "pilates", "crossfit", "martial",
    "boutique", "clothing", "jewelry", "furniture", "gift", "antique",
    "hotel", "motel", "inn", "bed_and_breakfast", "lodging", "vacation_rental",
    "law", "lawyer", "attorney", "accountant", "accounting", "tax", "bookkeep",
    "cleaning", "home_cleaning", "interior_design", "architect", "moving",
    "daycare", "day_care", "preschool", "tutor", "driving_school", "pet_groom",
    "pet_", "kennel", "dog_", "brewpub", "distillery", "cidery", "smoothie", "juice",
]
LOW_FIT = [
    "church", "cathedral", "religi", "mosque", "temple", "synagog", "worship",
    "government", "public_and_government", "city_hall", "courthouse", "dmv",
    "post_office", "police", "fire_station", "library", "school_district",
    "non_profit", "nonprofit", "charity", "community_services", "association",
    "political", "union", "embassy", "military", "prison", "cemetery",
    "utility", "power_plant", "water_treatment", "landfill", "recycling_center",
    "university", "college", "high_school", "elementary_school", "public_school",
    "hospital", "clinic_public", "atm", "parking", "rest_area", "toll",
]


def industry_fit(category):
    c = (category or "").lower()
    if not c:
        return "medium"
    for k in LOW_FIT:
        if k in c:
            return "low"
    for k in HIGH_FIT:
        if k in c:
            return "high"
    return "medium"


def pretty(cat):
    return (cat or "business").replace("_", " ").strip()


def outreach_score(tier, fit, has_phone, has_email, has_social, completeness, confidence):
    if not has_phone:
        return 0  # not callable -> not on the call sheet
    need = {"A": 50, "B": 42, "C": 14}.get(tier, 14)
    fitpts = {"high": 28, "medium": 8, "low": -25}[fit]
    reach = 12 + (8 if has_email else 0) + (3 if has_social else 0)  # phone is given
    extra = int((completeness or 0) * 0.08) + int(round((confidence or 0) * 4))
    return max(0, min(100, need + fitpts + reach + extra))


def why_warm(tier, fit, builder, has_email):
    bits = []
    if tier == "A":
        bits.append("No real website")
    elif tier == "B":
        bits.append(f"DIY site ({builder or 'builder'}) — upsell")
    else:
        bits.append("Has a site")
    if fit == "high":
        bits.append("high-fit niche")
    bits.append("has phone" + (" + email" if has_email else ""))
    return "; ".join(bits)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--top", type=int, default=300, help="rows in the xlsx Top Calls tab")
    args = ap.parse_args()

    db = sqlite3.connect(DB_PATH)
    have = {r[1] for r in db.execute("PRAGMA table_info(leads)")}
    for col, typ in (("outreach_score", "INTEGER"), ("industry_fit", "TEXT")):
        if col not in have:
            db.execute(f"ALTER TABLE leads ADD COLUMN {col} {typ}")

    rows = db.execute(
        "SELECT id, name, category, city, phone, phone_fmt, email, socials, "
        "website, tier, builder, completeness, confidence, pitch FROM leads"
    ).fetchall()

    enriched = []
    for (lid, name, cat, city, phone, pfmt, email, socials, website, tier,
         builder, comp, conf, pitch) in rows:
        fit = industry_fit(cat)
        has_phone = bool(phone and str(phone).strip())
        has_email = bool(email and str(email).strip())
        has_social = bool(socials and str(socials).strip())
        sc = outreach_score(tier, fit, has_phone, has_email, has_social, comp, conf)
        enriched.append({
            "id": lid, "name": name, "cat": cat, "city": city,
            "phone": pfmt or phone, "email": email, "tier": tier,
            "builder": builder, "fit": fit, "score": sc,
            "why": why_warm(tier, fit, builder, has_email),
            "pitch": pitch, "website": website,
        })

    # write scores back to leads
    db.execute("BEGIN")
    db.executemany("UPDATE leads SET outreach_score=?, industry_fit=? WHERE id=?",
                   [(e["score"], e["fit"], e["id"]) for e in enriched])
    db.commit()
    db.execute("CREATE INDEX IF NOT EXISTS idx_leads_outreach ON leads(outreach_score)")
    db.commit()

    # callable warm leads only, ranked
    warm = [e for e in enriched if e["score"] > 0 and e["fit"] != "low"]
    warm.sort(key=lambda e: (-e["score"], e["cat"] or "", e["name"] or ""))

    # ---- warm_leads.csv (everything callable & warm) ----
    import csv
    csv_path = os.path.join(DATA, "warm_leads.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["rank", "outreach_score", "tier", "industry_fit", "business",
                    "niche", "city", "phone", "email", "why_warm", "pitch",
                    "website", "id"])
        for i, e in enumerate(warm, 1):
            w.writerow([i, e["score"], e["tier"], e["fit"], e["name"], pretty(e["cat"]),
                        e["city"], e["phone"], e["email"] or "", e["why"],
                        e["pitch"], e["website"] or "", e["id"]])

    # ---- Monday_Call_Sheet.xlsx ----
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"  (openpyxl not installed; wrote {csv_path} only. "
              f"`pip install openpyxl` for the xlsx call sheet.)")
        db.close()
        return

    wb = openpyxl.Workbook()
    HEAD = PatternFill("solid", fgColor="1F4E78")
    HEADF = Font(color="FFFFFF", bold=True, size=11)
    HOT = PatternFill("solid", fgColor="C6EFCE")
    WARM = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def make_sheet(ws, data, title, note):
        ws.title = title
        ws.sheet_view.showGridLines = False
        ws["A1"] = note
        ws["A1"].font = Font(bold=True, size=12, color="1F4E78")
        headers = ["#", "Score", "Tier", "Business", "Niche", "City", "Phone",
                   "Email", "Why it's warm", "Suggested pitch (read/adapt)",
                   "Call outcome", "Follow-up notes"]
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=j, value=h)
            c.fill = HEAD; c.font = HEADF; c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        widths = [4, 6, 5, 30, 20, 14, 16, 26, 30, 60, 18, 24]
        for j, wdt in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = wdt
        for i, e in enumerate(data, 1):
            r = i + 3
            vals = [i, e["score"], e["tier"], e["name"], pretty(e["cat"]), e["city"],
                    e["phone"], e["email"] or "", e["why"], e["pitch"], "", ""]
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=j, value=v)
                c.border = border
                c.alignment = Alignment(vertical="top",
                                        wrap_text=(j in (9, 10, 12)))
            fill = HOT if e["score"] >= 80 else (WARM if e["score"] >= 65 else None)
            if fill:
                ws.cell(row=r, column=2).fill = fill
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{len(data)+3}"

    # Tab 1: Top Calls (all tiers, warmest first)
    make_sheet(wb.active, warm[:args.top], "Top Calls",
               f"MONDAY CALL SHEET — top {min(args.top, len(warm))} warmest leads "
               f"(of {len(warm):,} callable). Work top-down; green=hottest.")

    # Tab 2: Tier-B Upsell (they paid for a DIY site = they value web presence)
    bwarm = [e for e in warm if e["tier"] == "B"][:200]
    make_sheet(wb.create_sheet(), bwarm, "Tier-B Upsell",
               f"TIER-B UPSELL — {len(bwarm)} businesses on Wix/Weebly/etc. "
               f"They already pay for a site; pitch a faster custom one.")

    # Tab 3: By Niche (top niches among warm leads, with counts)
    ws3 = wb.create_sheet("By Niche")
    ws3.sheet_view.showGridLines = False
    ws3["A1"] = "WARM LEADS BY NICHE — where your hottest prospects are"
    ws3["A1"].font = Font(bold=True, size=12, color="1F4E78")
    for j, h in enumerate(["Niche", "Warm leads", "Avg score", "Tier-A (no site)"], 1):
        c = ws3.cell(row=3, column=j, value=h); c.fill = HEAD; c.font = HEADF
        c.alignment = Alignment(horizontal="center")
    ws3.column_dimensions["A"].width = 30
    for j in (2, 3, 4):
        ws3.column_dimensions[get_column_letter(j)].width = 14
    from collections import defaultdict
    agg = defaultdict(lambda: {"n": 0, "sum": 0, "a": 0})
    for e in warm:
        a = agg[pretty(e["cat"])]
        a["n"] += 1; a["sum"] += e["score"]; a["a"] += (1 if e["tier"] == "A" else 0)
    ranked = sorted(agg.items(), key=lambda kv: -kv[1]["n"])[:40]
    for i, (niche, a) in enumerate(ranked, 1):
        ws3.cell(row=i + 3, column=1, value=niche)
        ws3.cell(row=i + 3, column=2, value=a["n"])
        ws3.cell(row=i + 3, column=3, value=round(a["sum"] / a["n"], 1))
        ws3.cell(row=i + 3, column=4, value=a["a"])
    ws3.freeze_panes = "A4"

    xlsx_path = os.path.join(DATA, "Monday_Call_Sheet.xlsx")
    wb.save(xlsx_path)
    db.close()

    n_hot = sum(1 for e in warm if e["score"] >= 80)
    print(f"Wrote call sheet from {len(rows):,} leads.")
    print(f"  Callable warm leads: {len(warm):,}  (hot >=80: {n_hot:,})")
    print(f"  {xlsx_path}")
    print(f"    tabs: Top Calls ({min(args.top, len(warm))}), "
          f"Tier-B Upsell ({len(bwarm)}), By Niche")
    print(f"  {csv_path}  ({len(warm):,} rows)")


if __name__ == "__main__":
    main()
